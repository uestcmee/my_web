# coding:utf-8

import re

import chardet
import pandas as pd


def get_encoding(file):
    # 二进制方式读取，获取字节数据，检测类型
    with open(file, 'rb') as f:
        data = f.read()
        return chardet.detect(data)['encoding']


def pred_line(day='2021-01-05'):
    import akshare as ak
    # today=datetime.datetime.strptime('%Y-%m-%d',day)
    bond_china_yield_df = ak.bond_china_yield(start_date=day, end_date=day)
    today_line = bond_china_yield_df[bond_china_yield_df['曲线名称'] == '中债中短期票据收益率曲线(AAA)'].iloc[0]
    today_line = today_line.iloc[2:-1]
    today_line.index = [3 / 12, 6 / 12, 1, 3, 5, 7, 10]
    day_line = pd.Series(index=[0.25, 0.50] + [i for i in range(1, 11)], dtype=float)
    for one in day_line.index:
        if one in today_line.index:
            day_line.loc[one] = today_line[one]
        else:
            year0 = (today_line[today_line.index < one].index[-1])
            year1 = (today_line[today_line.index > one].index[0])
            rate0 = (today_line[today_line.index < one].iloc[-1])
            rate1 = (today_line[today_line.index > one].iloc[0])
            day_line.loc[one] = round((rate0 + (one - year0) / (year1 - year0) * (rate1 - rate0)), 4)
    return day_line


def plot_prepare(small_df, day='2021-01-05'):
    """
    :param small_df: 需要包括'收益率'和'day'（到期剩余天数）
    :param day: 当前日期
    :return:
    """
    day_line = pred_line(day=day)

    def pred_rate(year):  # 计算预测收益率
        if year < 1 and year < 0.5:
            y0 = 0.25
            y1 = 0.5
        elif year < 1 and year > 0.5:
            y0 = 0.5
            y1 = 1
        else:
            y0 = int(year)
            y1 = y0 + 1
        rate = day_line.loc[y0] + (year - y0) * (day_line.loc[y1] - day_line.loc[y0]) / (y1 - y0)
        return rate

    small_df = small_df[(small_df['year'] > 0.5) & (small_df['year'] < 5)]
    small_df['real_ytm'] = small_df['收益率'].apply(lambda x: float(re.findall('\d+\.\d+|(?<!-)\d+', x)[0]))
    small_df['pred'] = (small_df['year']).apply(lambda x: pred_rate(x))
    small_df['spread'] = small_df['real_ytm'] - small_df['pred']
    small_df = small_df.round(4)  # 小数位数
    small_df['year'] = small_df['year'].round(2)
    return small_df


def deal_process_func(file_name='2020年06月25日周四.txt'):
    encoding = get_encoding('./static/uploads/{}'.format(file_name))
    with open('./static/uploads/{}'.format(file_name), encoding=encoding) as f:
        text = f.read()
        f.close()
    fenlei = {'短融': [],
              '中票': [],
              '企业债': [],
              '其他': []}

    need_flag = True  # 目前标题是否为我们所需要的
    for i, line in enumerate(text.split('\n')):
        line = re.split('[\t ]+', line.strip())
        if len(line) == 1 and line != ['']:
            if re.search('|'.join(['短融', '中票', '企业债', '其他']), line[0]) != None:
                now_list = re.search('|'.join(['短融', '中票', '企业债', '其他']), line[0]).group()
                need_flag = True
            else:
                need_flag = False
        if not need_flag:
            continue  # 如果为非所需的数据，跳过
        if len(line) >= 5:
            year = line[0]
            try:
                rating = line[[re.search('AAA|(A\-1)', i) != None for i in line].index(True)]  # 找评级
                ytm = line[[re.search('\d+\.\d+|(?<!-)\d+', i) != None for i in line[3:]].index(True) + 3].replace('%',
                                                                                                                   '')  # 找ytm，需要区分不要找到短融评级
                name = line[[re.search('[^\x00-\xff]+', i) != None for i in line[1:]].index(True) + 1]  # 找名称
                fenlei[now_list].append([year, name, ytm, rating])
            except:
                pass
        else:
            pass

    fenlei_df = {i: pd.DataFrame(fenlei[i], columns='剩余期限 简称 收益率 评级'.split(' ')) for i in fenlei.keys()}
    def get_day(days: str):
        main_day = days.split('+')[0]
        if 'D' in main_day:
            real_day = float(re.findall('\d+\.\d+|\d+', main_day)[0])
        elif 'Y' in main_day:
            real_day = float(re.findall('\d+\.\d+|\d+', main_day)[0]) * 365
        else:
            real_day = main_day
        real_day = round(float(real_day), 1)
        return real_day

    for key in fenlei_df:  # 每一个债券类型
        df = fenlei_df[key]
        df['year'] = df['剩余期限'].apply(lambda x: get_day(x) / 365)
        # print(df)
        df.sort_values('year', inplace=True)
        df.index = [i for i in range(len(df))]
        df.to_excel('./data/BondDeal/{}.xlsx'.format(key))
        df = plot_prepare(df)
        fenlei_df[key] = df  #.drop('day', axis=1)  # 去掉辅助列

    # tot_df=pd.concat([df for df in fenlei_df.values()])
    output_excel(file_name, fenlei_df)
    return fenlei_df


def output_excel(file_name, fenlei_df):
    from shutil import copyfile
    file_name = file_name.split('.')[0]
    file_path = './static/download/'
    copyfile('{}页面模板.xlsx'.format(file_path), '{}{}.xlsx'.format(file_path, file_name))
    import openpyxl
    wb = openpyxl.load_workbook('{}{}.xlsx'.format(file_path, file_name))
    sht = wb['Sheet1']

    def get_excel_coor(now_line):
        page = int(now_line / 120)
        line = now_line % 60
        if now_line % 120 < 60:
            col = 'A'
        else:
            col = 'E'
        row = page * 62 + line + 3
        return str(col) + str(row)

    from openpyxl.styles import Alignment
    def small_title(sheet, coor, title):
        end_coor_dic = {'A': 'D', 'E': 'H'}
        hebing = coor + ':' + end_coor_dic[coor[0]] + coor[1:]
        sheet.merge_cells(hebing)
        sheet[coor] = title
        sheet[coor].alignment = Alignment(horizontal='center', vertical='center')
    small_title(sht, 'A3', '短融')

    # 取消合并所有单元格
    for page in range(6):
        page_range = 'A{}:H{}'.format(page * 62 + 3, page * 62 + 62)
        try:
            sht.unmerge_cells(page_range)
        except:
            pass

    sht['E1'] = file_name
    def write_area(sht, coor, write_df):
        if coor[0] == 'A':
            min_col = 1
            max_col = 4
        if coor[0] == 'E':
            min_col = 5
            max_col = 8
        min_row = int(coor[1:])
        max_row = min_row + len(write_df) - 1
        #     print(write_df)
        for i, row in enumerate(sht.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col, max_row=max_row)):
            for j, cell in enumerate(row):
                #             print(cell)
                try:
                    cell.value = write_df.iloc[i, j]
                except:
                    print(cell)

    now_line = 0

    for one in fenlei_df:
        # 输入标签
        coor = get_excel_coor(now_line)
        small_title(sht, coor, one)
        now_line += 1

        if len(fenlei_df[one]) == 0:
            continue
        line_left = 60 - now_line % 60  # 剩下多少行的空间
        small_df = fenlei_df[one]
        while len(small_df) > 0:
            #   print('还剩多少行:',line_left,'需要多少行:',len(small_df))
            coor = get_excel_coor(now_line)  # 计算新的坐标位置

            # 填充剩余位置
            write_df = small_df.iloc[:line_left]
            write_area(sht, coor, write_df)
            # 更新位置
            now_line += len(write_df)

            small_df = small_df.iloc[line_left:]
            line_left = 60 - now_line % 60  # 更新剩下多少行的空间
    wb.save('{}{}.xlsx'.format(file_path, file_name))


if __name__ == '__main__':
    deal_process_func('2021年01月05日周二.txt')
